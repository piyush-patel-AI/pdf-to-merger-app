from flask import Flask, request, jsonify, send_from_directory, send_file, render_template 
from flask_cors import CORS  # ✅ add this import
from werkzeug.utils import secure_filename
from PyPDF2 import PdfMerger
from PIL import Image
from docx import Document
from openpyxl import load_workbook, Workbook
from reportlab.pdfgen import canvas
import os, tempfile, threading, sys

# ✅ Create Flask app first
app = Flask(__name__)

# ✅ Then enable CORS (so Cloudflare link can access Flask)
CORS(app)


try:
    import pypandoc
except Exception:
    pypandoc = None


# ------------------------------------------------------------
# ✅ Path handling (PyInstaller compatible)
# ------------------------------------------------------------
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


# ------------------------------------------------------------
# Flask setup
# ------------------------------------------------------------
app = Flask(
    __name__,
    template_folder=resource_path("templates"),
    static_folder=resource_path("static")
)

UPLOADED_FILES = []
SAVE_AS_PDF = False
ALLOWED_EXTENSIONS = {'pdf', 'jpg', 'jpeg', 'png', 'docx', 'xlsx', 'txt'}
TEMP_DIR = tempfile.TemporaryDirectory()


@app.route('/')
def serve_frontend():
    return send_from_directory(resource_path('.'), 'index.html')


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route('/upload', methods=['POST'])
def upload_files():
    global UPLOADED_FILES, SAVE_AS_PDF
    UPLOADED_FILES.clear()

    files = request.files.getlist('files')
    SAVE_AS_PDF = request.form.get('save_as_pdf', 'false') == 'true'

    if not files:
        return jsonify({'error': 'No files uploaded.'}), 400

    ext = files[0].filename.rsplit('.', 1)[1].lower()
    for f in files:
        if not allowed_file(f.filename):
            return jsonify({'error': f'File not allowed: {f.filename}'}), 400
        if not f.filename.lower().endswith(ext):
            return jsonify({'error': 'All files must be of the same type.'}), 400

    try:
        for f in files:
            safe = secure_filename(f.filename)
            path = os.path.join(TEMP_DIR.name, safe)
            f.save(path)
            UPLOADED_FILES.append(path)

        return jsonify({'message': f'{len(UPLOADED_FILES)} files uploaded successfully.'}), 200
    except Exception as e:
        print("Upload error:", e)
        return jsonify({'error': 'Failed to save files.'}), 500


@app.route('/merge', methods=['POST'])
def merge_files():
    global UPLOADED_FILES, SAVE_AS_PDF

    if len(UPLOADED_FILES) < 2:
        return jsonify({'error': 'Upload at least two files first.'}), 400

    _, ext = os.path.splitext(UPLOADED_FILES[0])
    ext = ext[1:].lower()
    merged_ext = 'pdf' if SAVE_AS_PDF else ext
    merged_output = os.path.join(TEMP_DIR.name, f'merged_output.{merged_ext}')

    try:
        # --- PDF merge ---
        if ext == 'pdf':
            merger = PdfMerger()
            for f in UPLOADED_FILES:
                merger.append(f)
            merger.write(merged_output)
            merger.close()

        # --- Image merge ---
        elif ext in ['jpg', 'jpeg', 'png']:
            images = [Image.open(f).convert('RGB') for f in UPLOADED_FILES]
            if SAVE_AS_PDF:
                images[0].save(merged_output, save_all=True, append_images=images[1:], format="PDF")
            else:
                widths, heights = zip(*(img.size for img in images))
                total_height = sum(heights)
                max_width = max(widths)
                merged_img = Image.new('RGB', (max_width, total_height), (255, 255, 255))
                y_offset = 0
                for img in images:
                    merged_img.paste(img, (0, y_offset))
                    y_offset += img.height
                merged_img.save(merged_output)

        # --- DOCX merge ---
        elif ext == 'docx':
            merged_doc = Document()
            for f in UPLOADED_FILES:
                doc = Document(f)
                for element in doc.element.body:
                    merged_doc.element.body.append(element)
            merged_doc.save(merged_output)

        # --- XLSX merge ---
        elif ext == 'xlsx':
            merged_wb = Workbook()
            merged_ws = merged_wb.active
            current_row = 1
            for f in UPLOADED_FILES:
                wb = load_workbook(f, read_only=True)
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    for j, val in enumerate(row, start=1):
                        merged_ws.cell(row=current_row, column=j, value=val)
                    current_row += 1
            merged_wb.save(merged_output)

        # --- TXT merge ---
        elif ext == 'txt':
            merged_txt = os.path.join(TEMP_DIR.name, 'merged.txt')
            with open(merged_txt, 'w', encoding='utf-8') as out:
                for f in UPLOADED_FILES:
                    with open(f, 'r', encoding='utf-8') as inp:
                        out.write(inp.read() + "\n")

            if SAVE_AS_PDF:
                c = canvas.Canvas(merged_output)
                y = 800
                with open(merged_txt, 'r', encoding='utf-8') as mtxt:
                    for line in mtxt:
                        c.drawString(50, y, line.strip())
                        y -= 15
                        if y < 50:
                            c.showPage()
                            y = 800
                c.save()
            else:
                merged_output = merged_txt

        else:
            return jsonify({'error': 'Unsupported file type'}), 400

        # ✅ Return file directly for browser download
        return send_file(
            merged_output,
            as_attachment=True,
            download_name=f'merged_document.{merged_ext}'
        )

    except Exception as e:
        print(f"Merge error: {type(e).__name__}: {e}")
        return jsonify({'error': f'Merge failed: {e}'}), 500


if __name__ == '__main__':
    import atexit
    atexit.register(TEMP_DIR.cleanup)

    def run_flask():
        app.run(debug=False, host='0.0.0.0', port=5000, use_reloader=False)

    print("🌐 Server running on http://127.0.0.1:5000")
    run_flask()
